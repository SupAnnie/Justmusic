import cv2 as cv
import HandTrackingModule as HT
from openpyxl import Workbook,load_workbook
import os

cap = cv.VideoCapture(0)
detector = HT.handDetector(mode=False, maxHands = 1, detectionCon=0.5, trackCon=0.5)
gestures_list = []

num = -2

file = 2
while True:

    if cv.waitKey(33) == ord(' ') :
        print('已收集手势数量{}'.format(num))

        break
    success, img = cap.read()
    img = detector.findHands(img)
    data_list = detector.findPosition1(img, draw=False)
    # 摄像头镜像翻转
    if success == True:
        img = cv.flip(img, 180)

    cv.imshow('1', img)


    if len(data_list) != 0:
        print(num)
        data_list.insert(0, file)
        print(data_list)

        gestures_list.append(data_list)
        num = num + 1


path = 'figure3.xlsx'

if os.path.exists(path):
    gestures_book = load_workbook(path)#得到文件
    sheet = gestures_book.create_sheet(str(file))

else:
    gestures_book = Workbook()
    sheet = gestures_book.create_sheet(str(file))


for i in range(len(gestures_list)):
    for j in range(64):
        sheet.cell(i+1,j+1).value = gestures_list[i][j]

gestures_book.save(path)
print('file saved!')